from datetime import datetime
from flask import Blueprint, render_template, request, redirect, session, flash, jsonify, send_file
from app.models.database import get_db
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

log = Blueprint("log", __name__)


# =========================
# HALAMAN LOG ABSENSI
# =========================
@log.route("/log")
def halaman_log():
    if "login" not in session:
        return redirect("/")

    tanggal = request.args.get("tanggal")
    bulan   = request.args.get("bulan")
    tahun   = request.args.get("tahun")
    status  = request.args.get("status")

    if not tanggal and not (bulan and tahun):
        tanggal = datetime.now().strftime("%Y-%m-%d")

    return render_template(
        "log.html",
        tanggal=tanggal,
        bulan=bulan,
        tahun=tahun,
        status=status,
        today=datetime.now().strftime("%Y-%m-%d")
    )


# =========================
# DATA JSON UNTUK AUTO REFRESH
# =========================
@log.route("/log/data")
def log_data():
    tanggal = request.args.get("tanggal") or datetime.now().strftime("%Y-%m-%d")
    status  = request.args.get("status")   # 🔥 TAMBAHAN

    query = """
        SELECT 
            siswa.id as siswa_id,
            siswa.nama,
            siswa.nis,
            kelas.nama_kelas,
            COALESCE(absensi.status, 'alfa') as status,
            COALESCE(absensi.menit_telat, 0) as menit_telat,
            absensi.waktu_masuk,
            absensi.waktu_pulang,
            absensi.tanggal
        FROM siswa
        LEFT JOIN kelas ON siswa.kelas_id = kelas.id
        LEFT JOIN absensi ON siswa.id = absensi.siswa_id 
                         AND absensi.tanggal = ?
    """

    params = [tanggal]

    # =========================
    # FILTER STATUS
    # =========================
    if status:
        query += " WHERE COALESCE(absensi.status, 'alfa') = ?"
        params.append(status)

    query += " ORDER BY kelas.nama_kelas, siswa.nama"

    with get_db() as conn:
        rows = conn.execute(query, params).fetchall()

    return jsonify([dict(row) for row in rows])

# =========================
# AMBIL DAFTAR KELAS
# =========================
@log.route("/log/kelas")
def get_kelas_list():
    with get_db() as conn:
        rows = conn.execute("SELECT id, nama_kelas FROM kelas ORDER BY nama_kelas").fetchall()
    kelas_list = [{"id": row["id"], "nama_kelas": row["nama_kelas"]} for row in rows]
    return jsonify(kelas_list)


# =========================
# UPDATE ABSENSI
# =========================
@log.route("/update_absen/<int:siswa_id>", methods=["POST"])
def update_absen(siswa_id):
    if "login" not in session:
        return redirect("/")

    status = request.form.get("status")
    tanggal = request.form.get("tanggal") or datetime.now().strftime("%Y-%m-%d")

    if status not in ["hadir", "sakit", "izin", "alfa"]:
        flash("Status tidak valid", "danger")
        return redirect("/log")

    with get_db() as conn:
        c = conn.cursor()

        existing = c.execute(
            "SELECT id FROM absensi WHERE siswa_id = ? AND tanggal = ?",
            (siswa_id, tanggal)
        ).fetchone()

        if existing:
            c.execute("UPDATE absensi SET status = ? WHERE id = ?",
                     (status, existing["id"]))
        else:
            c.execute("""
                INSERT INTO absensi (siswa_id, tanggal, status, menit_telat)
                VALUES (?, ?, ?, 0)
            """, (siswa_id, tanggal, status))

        conn.commit()

    return redirect("/log")


# =========================
# DOWNLOAD ABSENSI
# =========================
@log.route("/log/download")
def download_absensi():
    if "login" not in session:
        return redirect("/")

    import calendar

    tipe      = request.args.get("tipe")
    tanggal   = request.args.get("tanggal")
    bulan     = request.args.get("bulan")
    tahun     = request.args.get("tahun")
    kelas_id  = request.args.get("kelas_id")

    if not tipe:
        flash("Pilih tipe download terlebih dahulu", "warning")
        return redirect("/log")

    # =========================
    # MAPPING STATUS (FIX)
    # =========================
    status_map = {
        "hadir": "H",
        "sakit": "S",
        "izin": "I",
        "alfa": "A"
    }

    with get_db() as conn:
        c = conn.cursor()

        # ====================== BULANAN ======================
        if tipe == "bulanan" and bulan and tahun:

            kelas_filter = " WHERE kelas.id = ?" if kelas_id else ""
            params = [kelas_id] if kelas_id else []

            siswa_rows = c.execute(f"""
                SELECT siswa.id, siswa.nis, siswa.nama, kelas.nama_kelas
                FROM siswa
                LEFT JOIN kelas ON siswa.kelas_id = kelas.id
                {kelas_filter}
                ORDER BY kelas.nama_kelas, siswa.nama
            """, params).fetchall()

            absen_rows = c.execute("""
                SELECT siswa_id, tanggal, status
                FROM absensi
                WHERE strftime('%m', tanggal) = ?
                AND strftime('%Y', tanggal) = ?
            """, (bulan.zfill(2), tahun)).fetchall()

            jumlah_hari = calendar.monthrange(int(tahun), int(bulan))[1]

            tanggal_list = [
                f"{tahun}-{bulan.zfill(2)}-{str(i).zfill(2)}"
                for i in range(1, jumlah_hari + 1)
            ]

            data = {}
            siswa_map = {}

            # INIT semua siswa = A
            for s in siswa_rows:
                key = (s["nis"], s["nama"], s["nama_kelas"] or "Tanpa Kelas")
                data[key] = {tgl: "A" for tgl in tanggal_list}
                siswa_map[s["id"]] = key

            # Isi absensi
            for row in absen_rows:
                key = siswa_map.get(row["siswa_id"])
                if key and row["tanggal"] in data[key]:
                    data[key][row["tanggal"]] = status_map.get(row["status"], "A")

            df = pd.DataFrame.from_dict(data, orient='index')
            df = df.reindex(columns=tanggal_list, fill_value="A")

            # ubah header jadi 01,02,03
            df.columns = [tgl[-2:] for tgl in tanggal_list]

            df.index.names = ['NIS', 'Nama Siswa', 'Kelas']
            df = df.reset_index()

            filename = f"Absensi_Bulanan_{bulan}-{tahun}.xlsx"

        # ====================== HARIAN ======================
        else:
            kelas_filter = " AND kelas.id = ?" if kelas_id else ""
            params = [tanggal, tanggal]

            if kelas_id:
                params.append(kelas_id)

            rows = c.execute(f"""
                SELECT 
                    siswa.nis,
                    siswa.nama,
                    kelas.nama_kelas,
                    COALESCE(absensi.tanggal, ?) as tanggal,
                    COALESCE(absensi.status, 'alfa') as status
                FROM siswa
                LEFT JOIN kelas ON siswa.kelas_id = kelas.id
                LEFT JOIN absensi ON siswa.id = absensi.siswa_id 
                                 AND absensi.tanggal = ?
                WHERE 1=1 {kelas_filter}
                ORDER BY kelas.nama_kelas, siswa.nama
            """, params).fetchall()

            data = []
            for row in rows:
                data.append({
                    "NIS": row["nis"],
                    "Nama Siswa": row["nama"],
                    "Kelas": row["nama_kelas"] or "-",
                    "Tanggal": row["tanggal"],
                    "Status": status_map.get(row["status"], "A"),
                })

            df = pd.DataFrame(data)
            filename = f"Absensi_Harian_{tanggal}.xlsx"

    # =========================
    # EXPORT EXCEL + FORMAT
    # =========================
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=4, sheet_name='Absensi')

        ws = writer.sheets['Absensi']

        # =========================
        # JUDUL
        # =========================
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws['A1'] = "FORMAT ABSENSI KEHADIRAN SISWA"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
        ws['A2'] = f"Bulan: {bulan or '-'} / Tahun: {tahun or '-'}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # =========================
        # HEADER STYLE
        # =========================
        from openpyxl.styles import PatternFill

        fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=5, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill

        # =========================
        # AUTO WIDTH
        # =========================
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2

        # =========================
        # BORDER + ALIGN
        # =========================
        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=5, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # =========================
        # WARNA STATUS (OPSIONAL KEREN)
        # =========================
        from openpyxl.styles import PatternFill

        for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
            for cell in row:
                if cell.value == "H":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")  # hijau
                elif cell.value == "A":
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")  # merah
                elif cell.value == "S":
                    cell.fill = PatternFill("solid", fgColor="BDD7EE")  # biru
                elif cell.value == "I":
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")  # kuning

    # =========================
    # RETURN FILE (FIX ERROR)
    # =========================
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )